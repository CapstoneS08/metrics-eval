import random
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

RANDOM_SEED = 42
random.seed(RANDOM_SEED)

ISSUE_FIELDS = [
    "issue",
    "customer",
    "created_by",
    "created_at",
    "to_inform",
    "assigned_to",
    "resolved_at",
    "status_to_close",
    "closed_at",
    "resolution_score",
    "comments",
]

META_COLS = [
    "message_id",
    "chat_id",
    "timestamp",
    "sender_name",
    "message_text",
]

def sample_df(df: pd.DataFrame, n: int) -> pd.DataFrame:
    if len(df) <= n:
        return df.copy()
    return df.sample(n=n, random_state=RANDOM_SEED).copy()

def _pack_output(row: pd.Series) -> str:
    packed = {k: row.get(k, "") for k in ISSUE_FIELDS}
    return str(packed)

def build_expert_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    # Ensure meta cols exist
    for c in META_COLS:
        if c not in out.columns:
            out[c] = ""

    out["output_json"] = out.apply(_pack_output, axis=1)

    # Rename message_text to input
    out = out.rename(columns={"message_text": "input"})

    base_cols = ["message_id", "chat_id", "timestamp", "sender_name", "input", "output_json"]
    out = out[base_cols]

    # Expert-only metrics (schema validity removed)
    out["Issue_Validity (Y/N)"] = ""
    out["Overall_Issue_Accuracy (Y/N)"] = ""

    # Field-level accuracy (Metric 3)
    for f in ISSUE_FIELDS:
        out[f"{f}_ok (Y/N/NA)"] = ""

    out["Notes"] = ""
    return out

def export_workbook(
    *,
    gpt_df: pd.DataFrame,
    claude_df: pd.DataFrame,
    out_path: Path,
    sample_n: int = 40,
):
    wb = Workbook()

    ws = wb.active
    ws.title = "ReadMe"

    lines = [
        "Issue Tracking Expert Validation",
        "",
        "Experts: Please fill ONLY these metrics:",
        "2) Issue Validity (%): Issue_Validity == Y",
        "3) Field-level Accuracy (%): each <field>_ok == Y (ignore NA)",
        "4) Overall Issue Accuracy (%): Overall_Issue_Accuracy == Y",
        "",
        "NOTE: Schema Validity (%) is computed automatically (not expert-filled).",
        "",
        "Instructions:",
        "- Use Y/N for Issue_Validity and Overall_Issue_Accuracy.",
        "- Use Y/N/NA for each field column (NA if not inferable from input).",
        "- Do not edit input/output columns.",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)

    gpt_sheet = build_expert_sheet(sample_df(gpt_df, sample_n))
    claude_sheet = build_expert_sheet(sample_df(claude_df, sample_n))

    ws_gpt = wb.create_sheet("GPT")
    for r in dataframe_to_rows(gpt_sheet, index=False, header=True):
        ws_gpt.append(r)

    ws_claude = wb.create_sheet("Claude")
    for r in dataframe_to_rows(claude_sheet, index=False, header=True):
        ws_claude.append(r)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print("Saved:", out_path)